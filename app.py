import os
import re
import pandas as pd
import pdfplumber
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template, abort
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import datetime
import time
from werkzeug.utils import secure_filename
import math
import re
import pdfplumber
import os
import re
import time
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter




app = Flask(__name__)
app.secret_key = 'your_secret_key'  

UPLOAD_FOLDER = 'uploads'
PROCESSED_FOLDER = 'processed_files'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER
for folder in [UPLOAD_FOLDER, PROCESSED_FOLDER]:
    os.makedirs(folder, exist_ok=True)

def get_db_connection():
    conn = sqlite3.connect('db.sqlite3')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()    



# --- Home Route ---
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

# --- Login Route ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ""
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['email'] = user['email']
            return redirect(url_for('dashboard'))
        else:
            error = "Invalid email or password."
            flash(error)
            return render_template('login.html', error=error)
    return render_template('login.html', error=error)

# --- Register Route ---
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')

    # Handles POST 
    data = request.get_json()
    name = data.get('username')
    email = data.get('email')
    password = data.get('password')

    # Validations
    if not name.replace(" ", "").isalpha():
        return {"success": False, "message": "Name must contain only letters."}

    if "jspmntc" not in email:
        return {"success": False, "message": "Email must contain 'jspmntc'."}

    if not re.match(r'^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$', password):
        return {"success": False, "message": "Password must be at least 6 characters with 1 capital, 1 number and 1 special character."}

    try:
        hashed_password = generate_password_hash(password)
        conn = get_db_connection()
        conn.execute('INSERT INTO users (name, email, password) VALUES (?, ?, ?)', (name, email, hashed_password))
        conn.commit()
        conn.close()
        return {"success": True, "message": "Registered successfully!"}
    except sqlite3.IntegrityError:
        return {"success": False, "message": "Email already registered."}

# --- Dashboard ---
@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash("Please log in first.")
        return redirect(url_for('login'))
    return render_template('dashboard.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template('forgotpass.html')

    #reset password
    data = request.get_json()
    email = data.get('email')
    new_password = data.get('new_password')

    if not email or not new_password:
        return {"success": False, "message": "All fields are required."}

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    if not user:
        conn.close()
        return {"success": False, "message": "Email not found."}

    # Replace old password with new hashed password
    hashed_password = generate_password_hash(new_password)
    conn.execute("UPDATE users SET password = ? WHERE email = ?", (hashed_password, email))
    conn.commit()
    conn.close()

    return {"success": True, "message": "Password updated successfully."}

# --- Logout Route ---
@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.')
    return redirect(url_for('login'))



def get_latest_uploaded_file():
    files = [f for f in os.listdir(UPLOAD_FOLDER) if f.endswith('.pdf')]
    if not files:
        return None
    files.sort(key=lambda x: os.path.getmtime(os.path.join(UPLOAD_FOLDER, x)), reverse=True)
    return os.path.join(UPLOAD_FOLDER, files[0])

def detect_mca_pattern(filepath):
    with pdfplumber.open(filepath) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        if not first_page_text:
            print(" ERROR: Unable to extract text from the first page!")
            return None
        first_page_lines = first_page_text.split("\n")[:20]
        first_page_text_cleaned = " ".join(first_page_lines).strip()
        print("\n======  DEBUG: First Page Extracted Text ======")
        print(first_page_text_cleaned)
        print("========================================")
        pattern_2020 = re.compile(r"MCA\)?\s*\(?\s*2020\s*Pattern", re.IGNORECASE)
        pattern_2024 = re.compile(r"MCA\)?\s*\(?\s*2024\s*Pattern", re.IGNORECASE)
        if pattern_2020.search(first_page_text_cleaned):
            print(" Detected MCA 2020 Pattern!")
            return "MCA 2020"
        elif pattern_2024.search(first_page_text_cleaned):
            print(" Detected MCA 2024 Pattern!")
            return "MCA 2024"
        else:
            print(" ERROR: No Pattern Detected!")
            print("👉 Extracted text:", first_page_text_cleaned)
            return None




@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"success": False, "message": "No file part"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"success": False, "message": "No selected file"}), 400

        
        original_filename = secure_filename(file.filename)
        unique_filename = f"{int(time.time())}_{original_filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        print(f"File saved to: {filepath}")

        # Process the uploaded file immediately.
        output_path = process_file(filepath)
        if not output_path:
            return jsonify({"success": False, "message": "Error processing file."}), 500

        processed_filename = os.path.basename(output_path)
        return jsonify({
            "success": True,
            "message": "File uploaded and processed successfully!",
            "file_name": processed_filename
        })
    except Exception as e:
        print("Exception in /upload route:", e)
        return jsonify({"success": False, "message": "Server error: " + str(e)}), 500

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    file_path = os.path.join(app.config['PROCESSED_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        abort(404)
# ---------- MCA 2020 Extraction Functions----------

# dynamic extraction.
EXPECTED_SUBHEADER_COUNT_2020 = 12

def extract_subheaders_2020(filepath):
    """
    Dynamically extract subheaders from pages starting from page 2.
    It scans for a candidate header line that splits into exactly 12 tokens.
    """
    
    with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages[1:]:
                for line in page.extract_text().split("\n"):
                    tokens = line.strip().split()
                    if 10 <= len(tokens) <= 14:
                        print(f"DEBUG header ({len(tokens)} tokens): {tokens}")
                    if len(tokens) == EXPECTED_SUBHEADER_COUNT_2020:
                        return tokens
    return None



def transform_dynamic_subheaders(tokens):
    """
    Transform the candidate dynamic subheader tokens into the desired list.
    """
    out = []
    try:
        out.append("INT_1")       
        out.append("INT_2")      
        out.append(tokens[2].upper()) 
        out.append(tokens[3].upper())  
        out.append(tokens[4].upper())  
        out.append(tokens[5].upper())  
        out.append("TOTAL")            
        out.append(tokens[7].upper())  
        out.append("ERN CRD")          
        out.append(tokens[10].upper()) 
        out.append("GRD PNT")          
        out.append("CRD PNT")          
    except Exception as e:
        print(f"Error in transform_dynamic_subheaders: {e}")
        return None

    if len(out) != 12:
        return None
    return out




def clean_token(token):
    return token.strip()

def preprocess_tokens(tokens):
    """
    Cleans extracted tokens by:
      1. Merging a leading "*" with the next token (if present),
      2. Removing isolated 'P' tokens,
      3. Removing a 'P' prefix if attached to a number,
      4. Replacing '---' with an empty string,
      5. Merging a numeric token with a following "FFF" token,
      6. Ensuring the final list has exactly 12 tokens (padding or truncating as needed).
    """
    merged_tokens = []
    skip_next = False
    for i, token in enumerate(tokens):
        if skip_next:
            skip_next = False
            continue
        if token.strip() == "*" and i + 1 < len(tokens):
            merged_tokens.append("*" + tokens[i + 1])
            skip_next = True
        else:
            merged_tokens.append(token)

    clean_tokens = []
    for token in merged_tokens:
        if token.strip() == "P":
            continue
        token = re.sub(r'^P\s*(\d+)', r'\1', token)  
        if token.strip() == "---":
            token = ""
        clean_tokens.append(token)

    merged2 = []
    i = 0
    while i < len(clean_tokens):
        tok = clean_tokens[i]
        if (i + 1 < len(clean_tokens)
            and clean_tokens[i+1].strip().upper() == "FFF"
            and any(ch.isdigit() for ch in tok)):
            merged2.append(tok.strip() + " FFF")
            i += 2
        else:
            merged2.append(tok)
            i += 1

    if len(merged2) < 12:
        merged2 += [""] * (12 - len(merged2))
    else:
        merged2 = merged2[:12]

    return merged2


def extract_subject_dict_2020(filepath):
    """
    Build subject_code → subject_name from Page 1 of the 2020 PDF.
    Matches lines like "CS101 Introduction to Programming".
    """
    subject_dict = {}
    with pdfplumber.open(filepath) as pdf:
        first_page = pdf.pages[0].extract_text() or ""
    for line in first_page.splitlines():
        line = line.strip()
        m = re.match(r'^([A-Z]{2}\d+[L]?|\d{3})\s+(.*)$', line)

        if m:
            code, name = m.groups()
            subject_dict[code] = name.strip()
    print("DEBUG 2020 Subject Dict:", subject_dict)
    return subject_dict


def process_subject_row_2020(current_student, subject_dict, subject_code, tokens, subheaders):
    
    #Clean tokens
    cleaned = preprocess_tokens(tokens)
    print(f"DEBUG 2020 [{subject_code}] after preprocess: {cleaned}")

    #truncate to match subheaders count
    n = len(subheaders)
    if len(cleaned) < n:
        cleaned += [""] * (n - len(cleaned))
    else:
        cleaned = cleaned[:n]
    print(f"DEBUG 2020 [{subject_code}] final tokens   : {cleaned}")

    #Map into current_student under keys "{SubjectName} {Header}"
    sem = current_student.get('Semester', '')             
    subj_name = subject_dict.get(subject_code, subject_code)
    for j, header in enumerate(subheaders):
        #subject column is prefixed with "Sem{N}"
        key = f"Sem{sem} {subj_name} {header}"
        current_student[key] = cleaned[j]

        print(f"  → {key}: {cleaned[j]}")


def extract_mca_2020_data(filepath, subheaders):
    subject_dict = extract_subject_dict_2020(filepath)
    if not subject_dict:
        return None

    student_data = []
    full_text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            if txt:
                full_text += txt + "\n"
    lines = full_text.splitlines()

    current_student = None
    current_subject_code = None
    current_subject_tokens = []

    # Pattern to catch SGPA + credits
    sgpa_pat = re.compile(
        r"Semester\s+SGPA\s*:\s*([\d.]+)\s+Credits\s+Earned/Total\s*:\s*(\S+)\s+Total\s+Credit\s+Points\s*:\s*(\S+)",
        re.IGNORECASE,
    )

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # SGPA line?
        if "Semester" in line and "SGPA" in line and current_student:
            m = sgpa_pat.search(line)
            if m:
                sgpa, cr_earned, cr_points = m.groups()
                if not sgpa.strip() or set(sgpa.strip()) == {"-"}:
                    sgpa = ""
                current_student["SGPA"] = sgpa
                current_student["Credits Earned"] = cr_earned
                current_student["Total Credit Points"] = cr_points
                print(f"DEBUG SGPA: {current_student['PRN']} → SGPA={sgpa}, Credits={cr_earned}, Points={cr_points}")
            continue

        #new student?
        if "PRN:" in line or line.upper().startswith("SEMESTER:"):
            # flush last subject
            if current_subject_code and current_subject_tokens and current_student:
                process_subject_row_2020(current_student, subject_dict,
                                        current_subject_code, current_subject_tokens, subheaders)
            current_subject_code = None
            current_subject_tokens = []

            if "PRN:" in line:
                if current_student:
                    student_data.append(current_student)
                info = re.search(r"PRN:(\d+)\s+SEAT NO.:(\d+)\s+NAME:(.*?)\s+Mother-", line)
                if info:
                    prn, seat_no, name = info.groups()
                else:
                    parts = line.split()
                    prn = parts[parts.index("PRN:") + 1] if "PRN:" in parts else ""
                    seat_no = parts[4] if len(parts) > 4 else ""
                    name = " ".join(parts[6:]) if len(parts) > 6 else ""
                current_student = {"PRN": prn, "Seat No": seat_no, "Name": name}
                print(f"DEBUG NEW STUDENT: PRN={prn}, Seat={seat_no}, Name={name}")
            continue

        # Subject row accumulation
        toks = line.split()
        code = toks[0].strip(":-.,").upper()
        if code in subject_dict:
            # flush previous
            if current_subject_code and current_subject_tokens and current_student:
                process_subject_row_2020(current_student, subject_dict,
                                        current_subject_code, current_subject_tokens, subheaders)
            current_subject_code = code
            current_subject_tokens = toks[1:]
            print(f"DEBUG NEW SUBJECT: {code} → tokens start {current_subject_tokens}")
        else:
            if current_subject_code is not None:
                current_subject_tokens.extend(toks)

    # final flush
    if current_subject_code and current_subject_tokens and current_student:
        process_subject_row_2020(current_student, subject_dict,
                                current_subject_code, current_subject_tokens, subheaders)
    if current_student:
        student_data.append(current_student)

    print(f"DEBUG 2020: Total students = {len(student_data)}")
    if student_data:
        print("DEBUG 2020 sample:", student_data[0])
    return student_data


def extract_mca_2020_data(filepath, subheaders=None):
    # 1) Fetch & transform 2020 subheaders
    dynamic_subs = extract_subheaders_2020(filepath)
    if not dynamic_subs or len(dynamic_subs) != EXPECTED_SUBHEADER_COUNT_2020:
        print("ERROR: Missing 12-token header for 2020.")
        return None
    subheaders = transform_dynamic_subheaders(dynamic_subs)
    if not subheaders:
        print("ERROR: transform_dynamic_subheaders failed.")
        return None

    # Build header
    raw_header_tokens = {tok.upper() for tok in dynamic_subs}
    common_headers = {
        "INT","INT_1","INT_2","EXT","PR","PJ","TOT","TOTAL",
        "CRD","ERN","GRD","PNT","GRD PNT","CRD PNT",
        "INT2","INT2_1","INT2_2","INT2 EXT","INT2 PR","INT2 PJ"
    }
    sem_headers = {f"SEM{n}" for n in range(1, 7)}
    header_skip = raw_header_tokens | common_headers | sem_headers

    # 2) Build subject dictionary
    subject_dict = {}
    with pdfplumber.open(filepath) as pdf:
        first_page = pdf.pages[0].extract_text() or ""
    for line in first_page.splitlines():
        parts = line.strip().split()
        if len(parts) < 2:
            continue
        code = parts[0].upper()
        if re.match(r'^[A-Z]{2}\d+[L]?$', code) or re.match(r'^\d{3}$', code):
            subject_dict[code] = " ".join(parts[1:]).strip()

    #Read all pages into a list of lines
    lines = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            lines.extend(txt.splitlines())

    # Precompile SGPA / credit regexes
    sgpa_re = re.compile(
        r"Semester\s+SGPA\s*:\s*([\d.]+)\s+Credits\s+Earned/Total\s*:\s*(\S+)\s+Total\s+Credit\s+Points\s*:\s*(\S+)",
        re.IGNORECASE
    )
    credit_re = re.compile(
        r"Credits\s+Earned/Total\s*:\s*(\S+)\s+Total\s+Credit\s+Points\s*:\s*(\S+)",
        re.IGNORECASE
    )

    student_data = []
    current_student = None
    current_subject_code = None
    current_subject_tokens = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        up = line.upper()
        if re.match(r'^SEM\d+\b', up) and 'SGPA' not in up and not up.startswith('SEMESTER:'):
            continue

        toks_up = [t.strip(':-.,').upper() for t in line.split()]
        
        if toks_up and all(tok in header_skip for tok in toks_up):
            continue
        tokens_raw = line.split()
        if len(tokens_raw) >= EXPECTED_SUBHEADER_COUNT_2020 + 1:
            hdr_seq = [tok.strip(':-.,').upper() for tok in tokens_raw[1:1+EXPECTED_SUBHEADER_COUNT_2020]]
            if hdr_seq == [h.upper() for h in dynamic_subs]:
                continue

        #PRN or SEMESTER header-flush pending subject, new student/semester
        if 'PRN:' in line or line.upper().startswith('SEMESTER:'):
            if current_subject_code and current_subject_tokens and current_student:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            current_subject_code = None
            current_subject_tokens = []

            if 'PRN:' in line:
            
                if current_student:
                    student_data.append(current_student)
                m = re.search(r"PRN:(\d+)\s+SEAT NO.:(\d+)\s+NAME:(.*?)\s+Mother-", line)
                if m:
                    prn, seat_no, name = m.groups()
                    
                    name = name.split('Mother')[0].strip()
                else:
                    parts = line.split()
                    prn     = parts[parts.index('PRN:') + 1]
                    seat_no = parts[4]
                    name    = " ".join(parts[6:])
                    name = name.split('Mother')[0].strip()
                current_student = {
                    'PRN': prn,
                    'Seat No': seat_no,
                    'Name': name,
                    'INT_count': 0,
                    'Semester': None
                }
            # assign semester
            if line.upper().startswith('SEMESTER:') and current_student:
                sem = line.split('SEMESTER:')[-1].strip().split()[0]
                current_student['Semester'] = sem
            continue

        # require valid current student + semester
        if not current_student or current_student.get('Semester') is None:
            continue

        if 'Semester' in line and 'SGPA' in line:
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
                current_subject_code = None
                current_subject_tokens = []
            sem = current_student['Semester']
            prefix = f"Sem{sem}"
            m1 = sgpa_re.search(line)
            if m1:
                    sgpa, ce, cp = m1.groups()
                    sgpa = '' if not sgpa.strip() or set(sgpa.strip()) == {'-'} else sgpa
                    current_student[f"{prefix} SGPA"] = sgpa
                    current_student[f"{prefix} Credits Earned"] = ce
                    current_student[f"{prefix} Total Credit Points"] = cp
            else:
                m2 = credit_re.search(line)
                if m2:
                        ce, cp = m2.groups()
                        current_student[f"{prefix} Credits Earned"] = ce
                        current_student[f"{prefix} Total Credit Points"] = cp
            continue
        if 'Credits Earned/Total' in line and 'SGPA' not in line:
            sem = current_student['Semester']
            prefix = f"Sem{sem}"
            m2 = credit_re.search(line)
            if m2:
                ce, cp = m2.groups()
                current_student[f"{prefix} Credits Earned"] = ce
                current_student[f"{prefix} Total Credit Points"] = cp
            continue
        if current_student is not None and "Credits Earned" not in current_student and "Credits Earned/Total" in line:
                credit_match = credit_re.search(line)
                if credit_match:
                    cr_earned, cr_points = credit_match.groups()
                    current_student["Credits Earned"] = cr_earned
                    current_student["Total Credit Points"] = cr_points
                    print(f"DEBUG (Credit): Assigned Credits Earned={cr_earned}, Total Credit Points={cr_points} to student {current_student.get('PRN','')}")
                    continue

        # tokenize for subjects
        parts = line.split()
        code = parts[0].strip(':-.,').upper()
        sem = current_student['Semester']

        # 4) Numeric code subjects (191, 192)
        if code.isdigit() and len(code) == 3 and code in subject_dict:
            subj_name = f"Sem{sem} {subject_dict[code]}"
            raw = parts[1:]
            processed = preprocess_tokens(raw)
            # pad/truncate
            processed = (processed + [""] * len(subheaders))[:len(subheaders)]
            # map fields
            mapping = {hdr: processed[i] for i, hdr in enumerate(subheaders)}
            if mapping['CRD'].isalpha():
                mapping['CRD'], mapping['GRD'] = mapping['GRD'], mapping['CRD']
            for hdr, val in mapping.items():
                current_student[f"{subj_name} {hdr}"] = val
            continue

        # 5) PRACTICAL
        is_prac = any(t.strip(':-.,').upper() == 'PRACTICAL' for t in parts)
        is_lab  = code.endswith('L') and code in subject_dict
        if is_prac or is_lab:
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            if is_prac:
                idx = next(i for i,t in enumerate(parts) if t.strip(':-.,').upper()=='PRACTICAL')
                raw = parts[idx+1:]
            else:
                raw = parts[1:]
            subj_code = f"PRACTICAL_{sem}"
            subject_dict[subj_code] = f"Sem{sem} PRACTICAL"
            process_subject_row_2020(
                current_student, subject_dict,
                subj_code, raw, subheaders
            )
            continue

        # 6) MINI PROJECT
        if code.startswith('ITC'):
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            raw = parts[1:]
            subj_code = f"MINIPROJ_{sem}"
            subject_dict[subj_code] = f"Sem{sem} MINI PROJECT"
            process_subject_row_2020(
                current_student, subject_dict,
                subj_code, raw, subheaders
            )
            continue

        # 7) CRD PNT
        if 'CRD PNT' in line.upper():
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            pos = line.upper().find('CRD PNT')
            raw = line[pos+len('CRD PNT'):].strip().split()
            subj_code = f"CRDPNT_{sem}"
            subject_dict[subj_code] = f"Sem{sem} CRD PNT"
            process_subject_row_2020(
                current_student, subject_dict,
                subj_code, raw, subheaders
            )
            continue

        # 8) INT
        if code == 'INT':
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            cnt = current_student['INT_count'] + 1
            current_student['INT_count'] = cnt
            raw = parts[2:]
            subj_code = f"INT_{sem}_{cnt}"
            subject_dict[subj_code] = 'INT' if cnt==1 else f"INT{cnt}"
            process_subject_row_2020(
                current_student, subject_dict,
                subj_code, raw, subheaders
            )
            continue

        # 9) Other subjects accumulator
        if code in subject_dict:
            if current_subject_code and current_subject_tokens:
                process_subject_row_2020(
                    current_student, subject_dict,
                    current_subject_code, current_subject_tokens, subheaders
                )
            current_subject_code = code
            current_subject_tokens = parts[1:]
        else:
            if current_subject_code:
                current_subject_tokens.extend(parts)

    # final flush
    if current_subject_code and current_subject_tokens and current_student:
        process_subject_row_2020(
            current_student, subject_dict,
            current_subject_code, current_subject_tokens, subheaders
        )
    if current_student:
        student_data.append(current_student)

    #Post-process: Overall Total, Percentage & ATKT per semester
    for student in student_data:
        # find semesters by SGPA keys
        sems = set()
        for key in student:
            m = re.match(r"Sem(\\d+) SGPA", key)
            if m:
                sems.add(m.group(1))
        for sem in sems:
            prefix = f"Sem{sem}"
            # Overall Total
            tot_keys = [k for k in student if k.startswith(f"{prefix} ") and k.endswith(" TOTAL")]
            total_vals = []
            for tk in tot_keys:
                try:
                    total_vals.append(float(student.get(tk) or 0))
                except:
                    total_vals.append(0.0)
            student[f"{prefix} Overall Total"] = sum(total_vals)
            # Percentage = SGPA * 9.5
            try:
                sgpa_val = float(student.get(f"{prefix} SGPA") or 0)
                student[f"{prefix} Percentage"] = sgpa_val * 9.5
            except:
                student[f"{prefix} Percentage"] = ""
            # ATKT = count of 'F' in GRD
            grd_keys = [k for k in student if k.startswith(f"{prefix} ") and k.endswith(" GRD")]
            atkt_count = sum(1 for gk in grd_keys if student.get(gk) == 'F')
            student[f"{prefix} ATKT"] = atkt_count

    return student_data














# ---------- Revised MCA 2024 Extraction Functions ----------

def extract_subject_dict_2024_old(filepath):
    """
    Extract the subject dictionary from Page 1 for MCA 2024.
    It looks for a header line containing both "Sub Code" and "Sub Name" and then uses a regex
    to split each subsequent line into subject code and subject name.
    Stops when a blank line or a line starting with "Min" or "Max" is encountered.
    """
    subject_dict = {}
    with pdfplumber.open(filepath) as pdf:
        first_page_text = pdf.pages[0].extract_text() or ""
    lines = first_page_text.splitlines()
    
    header_index = None
    for i, line in enumerate(lines):
        if "Sub Code" in line and "Sub Name" in line:
            header_index = i
            break
    if header_index is None:
        print("ERROR (2024): Could not find header line for subjects.")
        return None

    for line in lines[header_index+1:]:
        line = line.strip()
        if not line or line.startswith("Min") or line.startswith("Max"):
            break
        m = re.match(r"^(\S+)\s+(.*)$", line)
        if m:
            code, name = m.groups()
            subject_dict[code] = name
    print("✅ DEBUG (2024): Extracted Subject Dictionary:")
    print(subject_dict)
    return subject_dict

def extract_subheaders_2024_old(filepath):
    """
    Dynamically extract raw subheader tokens from the MCA 2024 PDF.
    This function scans the first 2 pages for a line with exactly 10 tokens.
    """
    EXPECTED_TOKEN_COUNT = 10
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages[:2]:
            text = pdf.pages[page.page_number - 1].extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if not line:
                    continue
                tokens = line.split()
                print(f"DEBUG (2024): Candidate header line: {tokens}")
                if len(tokens) == EXPECTED_TOKEN_COUNT:
                    print(f"DEBUG (2024): Found raw subheader tokens: {tokens}")
                    return tokens
    print("DEBUG (2024): No header line found with exactly 10 tokens.")
    return None

def transform_dynamic_subheaders_2024_old(raw_tokens):
    """
    Transform raw tokens (expected 10 tokens) into the final subheader list for MCA 2024.
    For example, if raw_tokens = ["INT", "EXT", "PR", "PJ", "TOT", "CRD", "ERN", "GRD", "GRD", "CRD"],
    the desired output is:
      ["INT", "EXT", "PR", "PJ", "TOTAL", "CRD", "ERN CRD", "GRD", "GRD PNT", "CRD PNT"]
    """
    EXPECTED_COUNT = 10
    if len(raw_tokens) < EXPECTED_COUNT:
        raw_tokens = raw_tokens + [""] * (EXPECTED_COUNT - len(raw_tokens))
    elif len(raw_tokens) > EXPECTED_COUNT:
        raw_tokens = raw_tokens[:EXPECTED_COUNT]
    try:
        out = []
        out.append(raw_tokens[0].upper())         
        out.append(raw_tokens[1].upper())         
        out.append(raw_tokens[2].upper())         
        out.append(raw_tokens[3].upper())         
        out.append("TOTAL")                       
        out.append(raw_tokens[5].upper())         
        out.append("ERN CRD")                     
        out.append(raw_tokens[8].upper() if raw_tokens[8] else "GRD")
        out.append("GRD PNT")                     
        out.append("CRD PNT")                     
    except Exception as e:
        print(f"Error in transform_dynamic_subheaders_2024: {e}")
        return None
    if len(out) != 10:
        print(f"Error: Final transformed subheaders count is {len(out)} instead of 10.")
        return None
    return out

def preprocess_tokens_2024_old(tokens):
    """
    Cleans extracted tokens for MCA 2024 subject rows:
      1. Merge a leading "*" with the next token to prevent column shifts.
      2. Remove standalone "*" tokens.
      3. Replace '---' with an empty string.
      4. Merge a token (that contains a digit) with the following "FFF" token.
      5. Remove tokens "191", "192", and "193".
      
    Returns the cleaned token list without enforcing a fixed length.
    The calling code should handle truncation or padding as needed.
    """
    
    merged_tokens = []
    skip_next = False
    for i, token in enumerate(tokens):
        if skip_next:
            skip_next = False
            continue
        if token.strip() == "*" and i + 1 < len(tokens):
            merged_tokens.append("*" + tokens[i+1])
            skip_next = True
        elif token.strip() != "*":
            merged_tokens.append(token)
    
    
    clean_tokens = ["" if token.strip() == "---" else token for token in merged_tokens]
    

    merged_tokens2 = []
    i = 0
    while i < len(clean_tokens):
        token = clean_tokens[i]
        if i < len(clean_tokens) - 1 and clean_tokens[i+1].strip().upper() == "FFF" and any(ch.isdigit() for ch in token):
            merged_tokens2.append(token.strip() + " FFF")
            i += 2  
        else:
            merged_tokens2.append(token)
            i += 1
    
    final_tokens = [tok for tok in merged_tokens2 if tok.strip() not in {"191", "192", "193"}]
    
    return final_tokens



def process_subject_row_2024(current_student, subject_dict, subject_code, tokens, subheaders, sem_prefix=""):

    """
    Process the accumulated tokens for a subject row:
      - Clean tokens using preprocess_tokens_2024.
      - Remove tokens "191", "192", and "193".
      - Remove any leading token that contains no digit (assuming it is an unwanted subject initial).
      - Pad or truncate the tokens to the expected number (based on subheaders).
      - Map tokens to the subheaders using the subject dictionary.
    """
    tokens = preprocess_tokens_2024_old(tokens)
    print(f"DEBUG: Tokens after preprocess_tokens_2024: {tokens}")
    

    tokens = [token for token in tokens if token not in ["191", "192", "193"]]
    print(f"DEBUG: Tokens after removing '191', '192', and '193': {tokens}")
    

    if tokens and not any(char.isdigit() for char in tokens[0]):
        print(f"DEBUG: Removing leading token '{tokens[0]}' (no digit found)")
        tokens = tokens[1:]
    
    
    expected_tokens = len(subheaders)
    if len(tokens) < expected_tokens:
        tokens += [""] * (expected_tokens - len(tokens))
    else:
        tokens = tokens[:expected_tokens]
    
    print(f"DEBUG: Final tokens used for marks: {tokens}")
    
    subj_name = subject_dict[subject_code]
    for j, header in enumerate(subheaders):
        key = f"{sem_prefix}{subj_name} {header}"
        current_student[key] = tokens[j]
        print(f"DEBUG: Mapping subject '{subj_name}' header '{header}' with token '{tokens[j]}'")


def extract_mca_2024_data(filepath, subheaders):
    """
    Extract MCA 2024 student data using logic adapted from MCA 2020.
      - Uses the subject dictionary (from Page 1) for subject code -> subject name mapping.
      - Builds keys as "{Subject Name} {Subheader}".
      - Allows rows with a minimum token count of 8 and pads them to the expected length.
    """
    subject_dict = extract_subject_dict_2024_old(filepath)
    if not subject_dict:
        return None
    print("====== Extracted Subject List (2024) ======")
    print(subject_dict)
    
    student_data = []
    full_text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            if txt:
                full_text += txt + "\n"
    lines = full_text.splitlines()
    current_student = None

    sgpa_pattern = re.compile(
        r"Semester\s+SGPA\s*:\s*([\d.]*)\s+Credits\s+Earned/Total\s*:\s*(\S+)\s+Total\s+Credit\s+Points\s*:\s*(\S+)",
        re.IGNORECASE,
    )
    # Additional credit pattern to extract credits if SGPA line is missing.
    credit_pattern = re.compile(
        r"Credits\s+Earned/Total\s*:\s*(\S+)\s+Total\s+Credit\s+Points\s*:\s*(\S+)",
        re.IGNORECASE,
    )

    # Variables to accumulate tokens for subject rows
    current_subject_tokens = []
    current_subject_code = None
    sem_prefix=""

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # detect semester header and set prefix
        if line.upper().startswith("SEMESTER:"):
            m = re.search(r"SEMESTER\s*:\s*(\d+)", line, re.IGNORECASE)
            if m:
                sem_prefix = f"Sem{m.group(1)} "
            continue

        # --- NEW SKIP BLOCK: ---
        parts = line.split()
        if parts and parts[0].strip(":-.,") in {"191", "192", "193"}:
            print(f"DEBUG (2024): Skipping line starting with unwanted token: {line}")
            continue
        # --- END SKIP BLOCK ---

        try:
        
            if "Semester" in line and "SGPA" in line and current_student:
                match = sgpa_pattern.search(line)
                if match and current_student is not None:
                    sgpa, cr_earned, cr_points = match.groups()
                    if not sgpa.strip() or set(sgpa.strip()) == {"-"}:
                        sgpa = ""
                    current_student[f"{sem_prefix}SGPA"]                = sgpa
                    current_student[f"{sem_prefix}Credits Earned"]      = cr_earned
                    current_student[f"{sem_prefix}Total Credit Points"] = cr_points
                    print(f"DEBUG (SGPA): Assigned SGPA={sgpa}, Credits Earned={cr_earned}, Total Credit Points={cr_points} to student {current_student.get('PRN','')}")
                    continue

            credit_key = f"{sem_prefix}Credits Earned"
            tcp_key    = f"{sem_prefix}Total Credit Points"
            if current_student is not None and credit_key not in current_student and "Credits Earned/Total" in line:
                credit_match = credit_pattern.search(line)
                if credit_match:
                    cr_earned, cr_points = credit_match.groups()
                    current_student[credit_key] = cr_earned
                    current_student[tcp_key]   = cr_points
                    print(f"DEBUG (Credit): Assigned Credits Earned={cr_earned}, Total Credit Points={cr_points} to student {current_student.get('PRN','')}")
                    continue

            # When encountering a PRN or SEMESTER line, flush any accumulated subject tokens.
            if "PRN:" in line or line.upper().startswith("SEMESTER:"):
                if current_subject_code and current_subject_tokens and current_student:
                   process_subject_row_2024(current_student, subject_dict, current_subject_code, current_subject_tokens, subheaders, sem_prefix)
                current_subject_tokens = []
                current_subject_code = None

                if "PRN:" in line:
                    if current_student:
                        student_data.append(current_student)
                    info_match = re.search(r"PRN:(\d+)\s+SEAT NO.:(\d+)\s+NAME:(.*?)\s+Mother-", line)
                    if info_match:
                        prn, seat_no, name = info_match.groups()
                    else:
                        parts = line.split()
                        prn = parts[parts.index("PRN:") + 1] if "PRN:" in parts else ""
                        seat_no = parts[4] if len(parts) > 4 else ""
                        name = " ".join(parts[6:]) if len(parts) > 6 else ""
                    current_student = {"PRN": prn, "Seat No": seat_no, "Name": name}
                    print(f"DEBUG: Started new student with PRN: {prn}, Seat No: {seat_no}, Name: {name}")
                    continue

            # Process subject rows:
            tokens = line.split()
            if not tokens:
                print("DEBUG (2024): Skipping empty token line")
                continue

            token0 = tokens[0].strip(":-.,").upper()
            print(f"DEBUG (2024): Processing student line. token0: '{token0}', tokens: {tokens}")

            if token0 in subject_dict:
                # New subject row detected; flush previous accumulation if needed.
                if current_subject_code and current_subject_tokens and current_student:
                    process_subject_row_2024(current_student, subject_dict, current_subject_code, current_subject_tokens, subheaders, sem_prefix)


                current_subject_code = token0
                current_subject_tokens = tokens[1:] 
            else:
                if current_subject_tokens is not None:
                    current_subject_tokens.extend(tokens)
        except Exception as e:
            print(f"Error processing line: {line} -- {e}")

    # Final flush for any remaining tokens.
    if current_subject_code and current_subject_tokens and current_student:
        process_subject_row_2024(current_student, subject_dict, current_subject_code, current_subject_tokens, subheaders, sem_prefix)

    
    if current_student:
        student_data.append(current_student)
    print(f"DEBUG (2024): Final student data count -> {len(student_data)}")
    if student_data:
        print(f"DEBUG (2024): Sample student -> {student_data[0]}")
    return student_data



# Helper functions
def safe_float(x):
    try:
        return float(x)
    except:
        return 0.0

def convert_to_number(val):
    try:
        return float(str(val).replace(',', '').strip())
    except:
        s = re.sub(r"[^\d\.]", "", str(val))
        return float(s) if s else 0.0

def calc_percentage(sgpa):
    t = str(sgpa).strip().lower()
    if t in ('', 'nan'):
        return ''
    try:
        return float(t) * 10 - 7.5
    except:
        return ''

def count_fail(row):
    return sum(
        1 for col, val in row.items()
        if col.upper().endswith('GRD') and str(val).strip().upper() == 'F'
    )

# Pattern-specific processors
def process_mca_2020(filepath):
    # Extract headers and raw data
    hdrs = extract_subheaders_2020(filepath) or []
    data = extract_mca_2020_data(filepath, hdrs)
    if not data:
        raise ValueError("No student data for MCA 2020.")

    df = pd.DataFrame(data)
    # Remove unneeded columns
    df.drop(columns=['INT_count', 'Semester'], errors='ignore', inplace=True)
    # Subject dictionary for toppers
    subj_dict = extract_subject_dict_2020(filepath) or {}
    return df, subj_dict, None

def process_mca_2024(filepath):
    subj_dict  = extract_subject_dict_2024_old(filepath) or {}
    raw_tokens = extract_subheaders_2024_old(filepath) or []
    hdrs       = transform_dynamic_subheaders_2024_old(raw_tokens) or []
    # Pass hdrs into the extractor
    data       = extract_mca_2024_data(filepath, hdrs)
    if not data:
        raise ValueError("No student data for MCA 2024.")

    df = pd.DataFrame(data)
    return df, subj_dict, hdrs




def finalize_results(df, subj_dict=None, hdrs=None, output_folder=None):
    """
    Finalize the DataFrame by computing per-semester metrics, reordering columns,
    building summaries and toppers, and writing to Excel.
    """
    # Ensure output folder exists
    output_folder = output_folder or app.config['PROCESSED_FOLDER']
    os.makedirs(output_folder, exist_ok=True)

    # Detect semesters
    cols = df.columns.tolist()
    sem_cols = [c for c in cols if re.match(r"^Sem\d+ SGPA$", c)]
    sem_numbers = sorted(int(re.search(r"Sem(\d+)", c).group(1)) for c in sem_cols)

    # Compute per-semester metrics
    for s in sem_numbers:
        base = f"Sem{s}"
        sg = f"{base} SGPA"
        pct = f"{base} Percentage"
        tcp = f"{base} Total Credit Points"
        tot = f"{base} Total"

        if sg in df:
            df[pct] = df[sg].apply(calc_percentage)
        df[tcp] = df.get(tcp, pd.Series(0, index=df.index)).map(convert_to_number)
        relevant = [c for c in cols if c.startswith(base) and 'TOTAL' in c.upper() and c != tcp]
        df[tot] = df[relevant].applymap(convert_to_number).sum(axis=1) if relevant else 0.0

    # Overall ATKT
    df['ATKT'] = df.apply(count_fail, axis=1)

    # Reorder final columns: core, summary, ATKT, then subjects in original PDF order
    core = [c for c in ('PRN', 'Seat No', 'Name') if c in df.columns]
    summary = []
    for s in sem_numbers:
        for suf in ('SGPA', 'Credits Earned', 'Total Credit Points', 'Percentage', 'Total'):
            col = f"Sem{s} {suf}"
            if col in df.columns:
                summary.append(col)
    atkt_col = ['ATKT'] if 'ATKT' in df.columns else []

    # Build subject columns list from subj_dict + hdrs to preserve PDF order
    # Build subject columns list, accounting for sem-prefix if present
    subject_cols = []
    if subj_dict and hdrs:
        # detect if subjects in df are prefixed like "Sem1 ADVANCED DBMS INT"
        prefix_mode = any(c.startswith('Sem') and any(subj in c for subj in subj_dict.values())
                          for c in df.columns)
        if prefix_mode:
            for s in sem_numbers:
                semp = f"Sem{s} "
                for code, subj_name in subj_dict.items():
                    for h in hdrs:
                        col = f"{semp}{subj_name} {h}"
                        if col in df.columns:
                            subject_cols.append(col)
        else:
            # fallback to old behavior
            for code, subj_name in subj_dict.items():
                for h in hdrs:
                    col = f"{subj_name} {h}"
                    if col in df.columns:
                        subject_cols.append(col)
    else:
        subject_cols = [c for c in df.columns if c not in core + summary + atkt_col]


    # Apply the final column order
    df = df[core + summary + atkt_col + subject_cols]

    # Subject summary sheet
    def build_subject_summary(df):
        grd_cols = [c for c in df.columns if re.match(r"^Sem\d+ .+ GRD$", c)]
        rows = []
        for col in grd_cols:
            sem, *parts, _ = col.split()
            sub = ' '.join(parts)
            grades = df[col].astype(str).str.strip().str.upper()
            total = len(grades)
            fails = (grades == 'F').sum()
            absent = (grades == '').sum()
            passed = total - fails - absent
            rows.append({
                'Semester': sem,
                'Subject': sub,
                'Total': total,
                'Passed': passed,
                'Failed': fails,
                'Absent': absent,
                'Pass %': (passed/total*100) if total else 0,
                'Fail %': (fails/total*100) if total else 0
            })
        return pd.DataFrame(rows).drop_duplicates(['Semester','Subject'])

    subject_summary = build_subject_summary(df)

    # Subject toppers sheet
    def build_subject_toppers(df, sem_numbers, subj_dict):
        rows = []
        for s in sem_numbers:
            prefix = f"Sem{s}"
            for code, sub in subj_dict.items():
                col = f"{prefix} {sub} TOTAL"
                if col not in df.columns:
                    continue
                tot_vals = pd.to_numeric(df[col], errors='coerce').fillna(0)
                maxv = tot_vals.max()
                names = df.loc[tot_vals == maxv, 'Name'].tolist()
                rows.append({
                    'Semester': prefix,
                    'Subject Code': code,
                    'Subject': sub,
                    'Topper(s)': ', '.join(names),
                    'Total': maxv
                })
        return pd.DataFrame(rows)

    subject_toppers = build_subject_toppers(df, sem_numbers, subj_dict or {})

    # SGPA toppers sheet
    def build_sgpa_top3(df, sem_numbers):
        rows = []
        for s in sem_numbers:
            col = f"Sem{s} SGPA"
            if col not in df.columns:
                continue
            sg = df[col].apply(safe_float)
            top_vals = sorted([v for v in sg.unique() if v > 0], reverse=True)[:3]
            for rank, val in enumerate(top_vals, start=1):
                names = df.loc[sg == val, 'Name'].tolist()
                rows.append({
                    'Semester': f"Sem{s}",
                    'Rank': rank,
                    'SGPA': val,
                    'Name(s)': ', '.join(names)
                })
        return pd.DataFrame(rows)

    sgpa_toppers = build_sgpa_top3(df, sem_numbers)

    # Write to Excel
    ts = int(time.time())
    fname = f"Master_Results_{ts}.xlsx"
    out = os.path.join(output_folder, fname)
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='All Semesters', index=False)
        subject_summary.to_excel(writer, sheet_name='Subject Summary', index=False)
        subject_toppers.to_excel(writer, sheet_name='Subject Toppers', index=False)
        sgpa_toppers.to_excel(writer, sheet_name='SGPA Toppers', index=False)
        for ws in writer.book.worksheets:
            ws.sheet_state = 'visible'

    format_excel(out, ['All Semesters','Subject Summary',
                   'Subject Toppers','SGPA Toppers'])        
    return out

# Main handler
def process_file(filepath):
    pattern = detect_mca_pattern(filepath)
    if not pattern:
        raise ValueError("Unknown MCA pattern")

    if pattern == 'MCA 2020':
        df, subj_dict, _ = process_mca_2020(filepath)
        hdrs = None 
    else:
        df, subj_dict, hdrs = process_mca_2024(filepath)

    return finalize_results(df, subj_dict, hdrs, output_folder=app.config['PROCESSED_FOLDER'])








def format_excel(filepath, sheets):
    """
    Apply a light-blue header, bold centered text, set all columns to width 15,
    AND collapse any two-row headers into one before styling.
    """
    wb = load_workbook(filepath)
    header_fill = PatternFill(start_color="ADD8E6",
                              end_color="ADD8E6",
                              fill_type="solid")
    bold_font   = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    for name in sheets:
        if name not in wb.sheetnames:
            print(f"⚠️  Sheet '{name}' not found—skipping")
            continue

        ws = wb[name]


        # ---- Style the single header row ----
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = bold_font
            cell.alignment = center

        # ---- Column widths ----
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(filepath)




if __name__ == '__main__':
    init_db()
    app.run(debug=True)